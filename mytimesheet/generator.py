"""Workbook generation for mytimesheet."""

from __future__ import annotations

import datetime as dt
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from .paths import beautify_path
    from .scanner import collect_modified_paths, get_work_date
except ImportError:
    from paths import beautify_path
    from scanner import collect_modified_paths, get_work_date


def scan_files(
    start_datetime: dt.datetime,
    end_datetime: dt.datetime,
    scan_dirs: list[Path],
    boundary: dt.time,
    scanner: str,
    quiet: bool,
) -> list[dict[str, object]]:
    if not quiet:
        print("Scanning file modifications...")
    paths = collect_modified_paths(scan_dirs, start_datetime, end_datetime, scanner, quiet)
    if not quiet:
        print(f"Found {len(paths)} files modified in timeframe.")

    records: list[dict[str, object]] = []
    for filepath in paths:
        try:
            timestamp = dt.datetime.fromtimestamp(filepath.stat().st_mtime)
        except OSError:
            continue
        if start_datetime <= timestamp <= end_datetime:
            folder_path = filepath.parent
            records.append(
                {
                    "timestamp": timestamp,
                    "work_date": get_work_date(timestamp, boundary),
                    "folder_path": folder_path,
                    "folder_name": folder_path.name or str(folder_path),
                }
            )

    if not quiet:
        print(f"Retained {len(records)} valid file modification records.")
    return records


def aggregate_activity(
    records: list[dict[str, object]], scan_dirs: list[Path]
) -> list[dict[str, object]]:
    grouped: dict[dt.date, dict[Path, list[dt.datetime]]] = defaultdict(lambda: defaultdict(list))
    for record in records:
        grouped[record["work_date"]][record["folder_path"]].append(record["timestamp"])

    aggregated: list[dict[str, object]] = []
    for work_date, folders in grouped.items():
        for folder_path, timestamps in folders.items():
            aggregated.append(
                {
                    "work_date": work_date,
                    "folder_name": folder_path.name or str(folder_path),
                    "folder_path": beautify_path(folder_path, scan_dirs),
                    "earliest_time": min(timestamps),
                    "latest_time": max(timestamps),
                    "edits_count": len(timestamps),
                }
            )
    aggregated.sort(key=lambda row: (row["work_date"], str(row["folder_name"])))
    return aggregated


def week_starts(start_date: dt.date, end_date: dt.date) -> list[dt.date]:
    first = start_date - dt.timedelta(days=start_date.weekday())
    weeks: list[dt.date] = []
    current = first
    while current <= end_date:
        weeks.append(current)
        current += dt.timedelta(days=7)
    return weeks


def generate_excel(
    aggregated_data: list[dict[str, object]],
    start_date: dt.date,
    end_date: dt.date,
    output_path: Path,
    title: str,
    boundary: dt.time,
) -> None:
    wb = openpyxl.Workbook()

    navy_color = "1B2D55"
    light_bg = "F5F7FA"
    muted_gray = "6C7A8D"
    border_gray = "D9D9D9"
    highlight_blue = "EAEFF8"
    font_family = "Segoe UI"

    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="E0E0E0")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    body_font = Font(name=font_family, size=11, color="000000")
    bold_body_font = Font(name=font_family, size=11, bold=True, color="000000")
    total_label_font = Font(name=font_family, size=11, bold=True, color=navy_color)
    gray_italics_font = Font(name=font_family, size=10, italic=True, color=muted_gray)

    navy_header_fill = PatternFill(start_color=navy_color, end_color=navy_color, fill_type="solid")
    zebra_fill = PatternFill(start_color=light_bg, end_color=light_bg, fill_type="solid")
    total_highlight_fill = PatternFill(
        start_color=highlight_blue, end_color=highlight_blue, fill_type="solid"
    )

    thin_border_side = Side(style="thin", color=border_gray)
    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )
    total_border = Border(
        top=Side(style="thin", color=navy_color),
        bottom=Side(style="double", color=navy_color),
    )

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.freeze_panes = "A4"
    ws_sum.views.sheetView[0].showGridLines = True

    ws_sum.merge_cells("A1:G1")
    ws_sum["A1"] = title
    ws_sum["A1"].font = title_font
    ws_sum["A1"].fill = navy_header_fill
    ws_sum["A1"].alignment = center_align
    ws_sum.row_dimensions[1].height = 30

    boundary_label = boundary.strftime("%I:%M %p").lstrip("0")
    ws_sum.merge_cells("A2:G2")
    ws_sum["A2"] = (
        f"Date range: {start_date:%Y-%m-%d} to {end_date:%Y-%m-%d} | "
        f"Workday boundary: {boundary_label}"
    )
    ws_sum["A2"].font = subtitle_font
    ws_sum["A2"].fill = navy_header_fill
    ws_sum["A2"].alignment = center_align
    ws_sum.row_dimensions[2].height = 18

    daily_headers = [
        ("Date", 1, center_align),
        ("Day", 2, center_align),
        ("Earliest Start Time", 3, center_align),
        ("Latest End Time", 4, center_align),
        ("Work Hours (Span)", 5, right_align),
        ("Is Workday", 6, center_align),
        ("Week Start", 7, center_align),
    ]
    for label, col, align in daily_headers:
        cell = ws_sum.cell(row=3, column=col, value=label)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align
        cell.border = thin_border
    ws_sum.row_dimensions[3].height = 24

    dates_list = [
        start_date + dt.timedelta(days=offset)
        for offset in range((end_date - start_date).days + 1)
    ]
    first_daily_row = 4
    for row, day in enumerate(dates_list, start=first_daily_row):
        ws_sum.row_dimensions[row].height = 20

        cell_date = ws_sum.cell(row=row, column=1)
        cell_date.value = dt.datetime.combine(day, dt.time.min)
        cell_date.number_format = "yyyy-mm-dd"
        cell_date.alignment = center_align
        cell_date.font = body_font

        cell_day = ws_sum.cell(row=row, column=2, value=f'=TEXT(A{row}, "ddd")')
        cell_day.alignment = center_align
        cell_day.font = body_font

        cell_start = ws_sum.cell(row=row, column=3)
        cell_start.value = (
            f'=IF(MAX(INDEX(\'Activity\'!$D$2:$D$5000 * '
            f'(\'Activity\'!$A$2:$A$5000 = A{row}), 0)) = 0, 0, '
            f'MIN(INDEX(\'Activity\'!$D$2:$D$5000 + '
            f'(\'Activity\'!$A$2:$A$5000 <> A{row}) * 99999, 0)))'
        )
        cell_start.number_format = "yyyy-mm-dd hh:mm AM/PM;;"
        cell_start.alignment = center_align
        cell_start.font = body_font

        cell_end = ws_sum.cell(row=row, column=4)
        cell_end.value = (
            f'=MAX(INDEX(\'Activity\'!$E$2:$E$5000 * '
            f'(\'Activity\'!$A$2:$A$5000 = A{row}), 0))'
        )
        cell_end.number_format = "yyyy-mm-dd hh:mm AM/PM;;"
        cell_end.alignment = center_align
        cell_end.font = body_font

        cell_hours = ws_sum.cell(row=row, column=5, value=f"=IF(C{row}=0, 0, (D{row}-C{row})*24)")
        cell_hours.number_format = "0.00"
        cell_hours.alignment = right_align
        cell_hours.font = body_font

        cell_workday = ws_sum.cell(row=row, column=6, value=f"=IF(E{row}>0, 1, 0)")
        cell_workday.number_format = "0"
        cell_workday.alignment = center_align
        cell_workday.font = body_font

        cell_week = ws_sum.cell(row=row, column=7)
        cell_week.value = f"=A{row}-WEEKDAY(A{row}, 2)+1"
        cell_week.number_format = "yyyy-mm-dd"
        cell_week.alignment = center_align
        cell_week.font = body_font

        for col in range(1, 8):
            cell = ws_sum.cell(row=row, column=col)
            cell.border = thin_border
            if row % 2 == 1:
                cell.fill = zebra_fill

    last_daily_row = first_daily_row + len(dates_list) - 1

    weekly_headers = [
        ("Week Start", 9, center_align),
        ("Week End", 10, center_align),
        ("Total Hours", 11, right_align),
        ("Work Days", 12, center_align),
    ]
    for label, col, align in weekly_headers:
        cell = ws_sum.cell(row=3, column=col, value=label)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align
        cell.border = thin_border

    weekly_first_row = 4
    for row, wk_start in enumerate(week_starts(start_date, end_date), start=weekly_first_row):
        wk_end = wk_start + dt.timedelta(days=6)
        ws_sum.cell(row=row, column=9, value=dt.datetime.combine(wk_start, dt.time.min))
        ws_sum.cell(row=row, column=10, value=dt.datetime.combine(wk_end, dt.time.min))
        ws_sum.cell(row=row, column=11).value = (
            f'=SUMIFS($E${first_daily_row}:$E${last_daily_row},'
            f'$G${first_daily_row}:$G${last_daily_row},I{row})'
        )
        ws_sum.cell(row=row, column=12).value = (
            f'=SUMIFS($F${first_daily_row}:$F${last_daily_row},'
            f'$G${first_daily_row}:$G${last_daily_row},I{row})'
        )
        for col in range(9, 13):
            cell = ws_sum.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = bold_body_font if col in {9, 10} else body_font
            cell.alignment = right_align if col == 11 else center_align
            if col in {9, 10}:
                cell.number_format = "yyyy-mm-dd"
            elif col == 11:
                cell.number_format = "0.00"
            else:
                cell.number_format = "0"
            if row % 2 == 1:
                cell.fill = zebra_fill

    total_row = weekly_first_row + len(week_starts(start_date, end_date))
    ws_sum.cell(row=total_row, column=9, value="Grand Total")
    ws_sum.merge_cells(start_row=total_row, start_column=9, end_row=total_row, end_column=10)
    ws_sum.cell(row=total_row, column=11).value = f"=SUM(K{weekly_first_row}:K{total_row - 1})"
    ws_sum.cell(row=total_row, column=12).value = f"=SUM(L{weekly_first_row}:L{total_row - 1})"
    for col in range(9, 13):
        cell = ws_sum.cell(row=total_row, column=col)
        cell.font = total_label_font if col in {9, 10} else bold_body_font
        cell.fill = total_highlight_fill
        cell.border = total_border
        cell.alignment = right_align if col == 11 else center_align
        cell.number_format = "0.00" if col == 11 else "0"

    note_row = total_row + 2
    ws_sum.merge_cells(start_row=note_row, start_column=9, end_row=note_row + 2, end_column=12)
    note_cell = ws_sum.cell(row=note_row, column=9)
    note_cell.value = (
        "Note:\nTotal work hours estimate the elapsed span from first file activity "
        "to last file activity. The Activity sheet is folder-level evidence, not a "
        "literal task timer."
    )
    note_cell.font = gray_italics_font
    note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws_act = wb.create_sheet(title="Activity")
    ws_act.freeze_panes = "A2"
    ws_act.views.sheetView[0].showGridLines = True

    act_headers = [
        ("Work Date", 1, center_align),
        ("Folder Name", 2, left_align),
        ("Folder Path", 3, left_align),
        ("Earliest Action", 4, center_align),
        ("Latest Action", 5, center_align),
        ("File Modifications", 6, center_align),
    ]
    for label, col, align in act_headers:
        cell = ws_act.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align
        cell.border = thin_border
    ws_act.row_dimensions[1].height = 24

    for row, record in enumerate(aggregated_data, start=2):
        ws_act.cell(row=row, column=1, value=dt.datetime.combine(record["work_date"], dt.time.min))
        ws_act.cell(row=row, column=2, value=record["folder_name"])
        ws_act.cell(row=row, column=3, value=record["folder_path"])
        ws_act.cell(row=row, column=4, value=record["earliest_time"])
        ws_act.cell(row=row, column=5, value=record["latest_time"])
        ws_act.cell(row=row, column=6, value=record["edits_count"])
        for col in range(1, 7):
            cell = ws_act.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = body_font
            cell.alignment = left_align if col in {2, 3} else center_align
            if col == 1:
                cell.number_format = "yyyy-mm-dd"
            elif col in {4, 5}:
                cell.number_format = "yyyy-mm-dd hh:mm AM/PM"
            elif col == 6:
                cell.number_format = "0"
            if row % 2 == 1:
                cell.fill = zebra_fill

    fold_headers = [("Folder Name Summary", 8, left_align), ("Total Edits", 9, center_align)]
    for label, col, align in fold_headers:
        cell = ws_act.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align
        cell.border = thin_border

    unique_folders = sorted({str(row["folder_name"]) for row in aggregated_data})
    for row, folder in enumerate(unique_folders, start=2):
        ws_act.cell(row=row, column=8, value=folder)
        ws_act.cell(row=row, column=9).value = f"=SUMIF($B$2:$B$5000,H{row},$F$2:$F$5000)"
        for col in range(8, 10):
            cell = ws_act.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = bold_body_font if col == 8 else body_font
            cell.alignment = left_align if col == 8 else center_align
            cell.number_format = "0" if col == 9 else "General"
            if row % 2 == 1:
                cell.fill = zebra_fill

    folder_total_row = max(2, len(unique_folders) + 2)
    ws_act.cell(row=folder_total_row, column=8, value="Grand Total")
    ws_act.cell(row=folder_total_row, column=9).value = (
        f"=SUM(I2:I{folder_total_row - 1})" if unique_folders else "=0"
    )
    for col in range(8, 10):
        cell = ws_act.cell(row=folder_total_row, column=col)
        cell.font = total_label_font if col == 8 else bold_body_font
        cell.fill = total_highlight_fill
        cell.border = total_border
        cell.alignment = left_align if col == 8 else center_align
        cell.number_format = "0" if col == 9 else "General"

    widths = {
        "Summary": {"A": 13, "B": 11, "C": 24, "D": 24, "E": 18, "F": 12, "G": 13, "I": 13, "J": 13, "K": 14, "L": 12},
        "Activity": {"A": 13, "B": 24, "C": 58, "D": 24, "E": 24, "F": 18, "H": 30, "I": 12},
    }
    for ws in [ws_sum, ws_act]:
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            if letter in widths.get(ws.title, {}):
                ws.column_dimensions[letter].width = widths[ws.title][letter]
            else:
                max_len = max(
                    (len(str(cell.value)) for cell in col_cells if cell.value and not str(cell.value).startswith("=")),
                    default=8,
                )
                ws.column_dimensions[letter].width = min(max(max_len + 3, 10), 40)

    wb.save(output_path)


def generate_timesheet(
    start_date: dt.date,
    end_date: dt.date,
    output_path: Path,
    scan_dirs: list[Path],
    boundary: dt.time = dt.time(2, 0),
    scanner: str = "auto",
    title: str = "WORKDAY & TIMESHEET SUMMARY",
    quiet: bool = False,
) -> None:
    start_dt = dt.datetime.combine(start_date, boundary)
    end_dt = dt.datetime.combine(end_date + dt.timedelta(days=1), boundary)

    records = scan_files(start_dt, end_dt, scan_dirs, boundary, scanner, quiet)
    aggregated = aggregate_activity(records, scan_dirs)
    generate_excel(aggregated, start_date, end_date, output_path, title, boundary)
