#!/usr/bin/env python3
"""
Texas 2036 Time Sheet Generator
Author: Antigravity AI Pair Programmer
For: Eric Booth, Senior Research Analyst, Texas 2036

Description:
  This script scans specified directories for file modification activity
  between specified dates. It shifts activity times by 2 hours (workdays extend 
  up to 2:00 AM the next calendar day and count towards the previous day). 
  It aggregates this activity by work date and folder, and outputs a 
  professionally designed, formula-driven Excel spreadsheet with Texas 2036 branding.
  
  Note: This version uses legacy-compatible INDEX/MAX/MIN array-equivalent formulas 
  instead of MINIFS/MAXIFS to ensure 100% compatibility across all versions of 
  Excel (including Excel 2016 and earlier) and Google Sheets.
"""

import os
import sys
import datetime
import subprocess
import argparse
from collections import defaultdict

# Try importing openpyxl and report error if missing
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: The 'openpyxl' library is required to run this script.")
    print("Please install it using: pip install openpyxl")
    sys.exit(1)

# Default configuration
DEFAULT_START_DATE = "2026-06-01"
DEFAULT_END_DATE = "2026-06-18"
DEFAULT_OUTPUT_FILE = "timesheet.xlsx"

# Folders to scan
TARGET_DIRS = [
    "/Users/ericbooth/Library/CloudStorage/GoogleDrive-eric.booth@texas2036.org/My Drive",
    "/Users/ericbooth/Library/CloudStorage/GoogleDrive-eric.booth@texas2036.org/Shared drives",
    "/Users/ericbooth/Documents",
    "/Users/ericbooth/Desktop"
]

# Excluded paths and extensions
EXCLUDE_SUBSTRINGS = [
    "/.git/", "/.venv/", "/venv/", "/node_modules/", "/.cache/", 
    "/__pycache__/", "/.claude/", "/.gemini/", "/.codex/", "/.trash/", 
    "/.metadata/", "/.vscode/", "/.settings/", "tmp/", "temp/"
]

# Exclude standard file names
EXCLUDE_FILENAMES = [
    ".ds_store", "desktop.ini", "thumbs.db"
]

def beautify_path(path):
    """Clean absolute path to be relative and user-friendly."""
    gdrive_prefix = "/Users/ericbooth/Library/CloudStorage/GoogleDrive-eric.booth@texas2036.org"
    home_prefix = "/Users/ericbooth"
    
    if path.startswith(gdrive_prefix):
        rel = os.path.relpath(path, gdrive_prefix)
        return rel
    elif path.startswith(home_prefix):
        rel = os.path.relpath(path, home_prefix)
        return f"~/{rel}"
    return path

def get_work_date(dt):
    """
    Determine the work date for a timestamp.
    Workday extends past midnight up to 2:00 AM (counts on the previous calendar day).
    """
    if dt.time() < datetime.time(2, 0, 0):
        return (dt - datetime.timedelta(days=1)).date()
    return dt.date()

def scan_files(start_datetime, end_datetime):
    """Scan directories using Spotlight mdfind for efficiency."""
    print("Scanning file modifications...")
    files_list = []
    
    start_str = start_datetime.strftime("%Y-%m-%dT%H:%M:%S")
    end_str = end_datetime.strftime("%Y-%m-%dT%H:%M:%S")
    
    for directory in TARGET_DIRS:
        if not os.path.exists(directory):
            print(f"  Directory not found (skipping): {directory}")
            continue
            
        print(f"  Querying Spotlight for: {directory}")
        query = f"kMDItemContentModificationDate >= $time.iso({start_str}) && kMDItemContentModificationDate <= $time.iso({end_str})"
        cmd = ["mdfind", "-onlyin", directory, query]
        
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, check=True)
            paths = result.stdout.splitlines()
            for path in paths:
                if os.path.isfile(path):
                    files_list.append(path)
        except Exception as e:
            print(f"  Warning: failed to query directory {directory}: {e}")
            
    # Remove duplicates
    files_list = list(set(files_list))
    print(f"Found {len(files_list)} files modified in timeframe from Spotlight.")
    
    # Extract modification times and filter
    activity_records = []
    for filepath in files_list:
        lower_path = filepath.lower()
        filename = os.path.basename(filepath).lower()
        
        if filename in EXCLUDE_FILENAMES:
            continue
        if any(exclude in lower_path for exclude in EXCLUDE_SUBSTRINGS):
            continue
            
        try:
            mtime = os.path.getmtime(filepath)
            dt = datetime.datetime.fromtimestamp(mtime)
            
            if start_datetime <= dt <= end_datetime:
                work_date = get_work_date(dt)
                folder_path = os.path.dirname(filepath)
                activity_records.append({
                    'timestamp': dt,
                    'work_date': work_date,
                    'folder_path': folder_path,
                    'folder_name': os.path.basename(folder_path) or folder_path
                })
        except Exception:
            continue
            
    print(f"Retained {len(activity_records)} valid file modification records.")
    return activity_records

def aggregate_activity(records):
    """Group activity by work date and folder to find earliest/latest times and edit counts."""
    grouped = defaultdict(lambda: defaultdict(list))
    
    for r in records:
        grouped[r['work_date']][r['folder_path']].append(r['timestamp'])
        
    aggregated = []
    for w_date, folders in grouped.items():
        for f_path, tstamps in folders.items():
            earliest = min(tstamps)
            latest = max(tstamps)
            folder_name = os.path.basename(f_path)
            pretty_path = beautify_path(f_path)
            
            aggregated.append({
                'work_date': w_date,
                'folder_name': folder_name,
                'folder_path': pretty_path,
                'earliest_time': earliest,
                'latest_time': latest,
                'edits_count': len(tstamps)
            })
            
    # Sort by date, then folder name
    aggregated.sort(key=lambda x: (x['work_date'], x['folder_name']))
    return aggregated

def generate_excel(aggregated_data, start_date, end_date, output_filename):
    """Generate a beautifully formatted Excel workbook using openpyxl."""
    print(f"Generating spreadsheet: {output_filename}")
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Styles Definition (Texas 2036 Palette & Tufte Guidelines)
    # ----------------------------------------------------
    navy_color = "1B2D55"
    light_bg = "F5F7FA"
    muted_gray = "6C7A8D"
    border_gray = "D9D9D9"
    highlight_blue = "EAEFF8"
    
    font_family = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="E0E0E0")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    body_font = Font(name=font_family, size=11, color="000000")
    bold_body_font = Font(name=font_family, size=11, bold=True, color="000000")
    total_label_font = Font(name=font_family, size=11, bold=True, color=navy_color)
    gray_italics_font = Font(name=font_family, size=10, italic=True, color=muted_gray)
    
    # Fills
    navy_header_fill = PatternFill(start_color=navy_color, end_color=navy_color, fill_type="solid")
    zebra_fill = PatternFill(start_color=light_bg, end_color=light_bg, fill_type="solid")
    total_highlight_fill = PatternFill(start_color=highlight_blue, end_color=highlight_blue, fill_type="solid")
    
    # Borders
    thin_border_side = Side(style='thin', color=border_gray)
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    top_thin_side = Side(style='thin', color=navy_color)
    double_bottom_side = Side(style='double', color=navy_color)
    total_border = Border(top=top_thin_side, bottom=double_bottom_side)
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    # ----------------------------------------------------
    # Sheet 1: Summary
    # ----------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_sum.merge_cells("A1:G1")
    title_cell = ws_sum["A1"]
    title_cell.value = "TEXAS 2036 WORKDAY & TIMESHEET SUMMARY"
    title_cell.font = title_font
    title_cell.fill = navy_header_fill
    title_cell.alignment = center_align
    ws_sum.row_dimensions[1].height = 30
    
    ws_sum.merge_cells("A2:G2")
    sub_cell = ws_sum["A2"]
    sub_cell.value = f"Date range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')} | Workday boundary: 2:00 AM"
    sub_cell.font = subtitle_font
    sub_cell.fill = navy_header_fill
    sub_cell.alignment = center_align
    ws_sum.row_dimensions[2].height = 18
    
    # Headers - Daily Log (A3:G3)
    daily_headers = [
        ("Date", "A3", center_align),
        ("Day", "B3", center_align),
        ("Earliest Start Time", "C3", center_align),
        ("Latest End Time", "D3", center_align),
        ("Work Hours (Span)", "E3", right_align),
        ("Is Workday", "F3", center_align),
        ("Week #", "G3", center_align)
    ]
    
    ws_sum.row_dimensions[3].height = 24
    for label, cell_ref, align in daily_headers:
        cell = ws_sum[cell_ref]
        cell.value = label
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align
        cell.border = thin_border
        
    # Generate dates list
    delta = end_date - start_date
    dates_list = [start_date + datetime.timedelta(days=i) for i in range(delta.days + 1)]
    
    # Populate Daily Log Rows (A4:G[N])
    current_row = 4
    for dt in dates_list:
        ws_sum.row_dimensions[current_row].height = 20
        
        # Col A: Date (stored as native datetime at midnight to match Activity Sheet exactly)
        cell_date = ws_sum.cell(row=current_row, column=1)
        cell_date.value = datetime.datetime.combine(dt, datetime.time.min)
        cell_date.number_format = 'yyyy-mm-dd'
        cell_date.alignment = center_align
        cell_date.font = body_font
        
        # Col B: Day
        cell_day = ws_sum.cell(row=current_row, column=2)
        cell_day.value = f'=TEXT(A{current_row}, "ddd")'
        cell_day.alignment = center_align
        cell_day.font = body_font
        
        # Col C: Earliest Start (Array-equivalent legacy formula using INDEX to avoid MINIFS/MAXIFS)
        # If MAX(INDEX(matches)) = 0, there are no matches, so it returns 0.
        # Otherwise, MIN(INDEX(matching_values + non_matching_values * 99999)) finds the earliest matching datetime.
        cell_start = ws_sum.cell(row=current_row, column=3)
        cell_start.value = (
            f'=IF(MAX(INDEX(Activity!$D$2:$D$5000 * (Activity!$A$2:$A$5000 = A{current_row}), 0)) = 0, 0, '
            f'MIN(INDEX(Activity!$D$2:$D$5000 + (Activity!$A$2:$A$5000 <> A{current_row}) * 99999, 0)))'
        )
        cell_start.number_format = 'yyyy-mm-dd hh:mm AM/PM;;'
        cell_start.alignment = center_align
        cell_start.font = body_font
        
        # Col D: Latest End (Array-equivalent MAX(INDEX(...)) to avoid MAXIFS)
        cell_end = ws_sum.cell(row=current_row, column=4)
        cell_end.value = f'=MAX(INDEX(Activity!$E$2:$E$5000 * (Activity!$A$2:$A$5000 = A{current_row}), 0))'
        cell_end.number_format = 'yyyy-mm-dd hh:mm AM/PM;;'
        cell_end.alignment = center_align
        cell_end.font = body_font
        
        # Col E: Work Hours Span (Evaluates to 0 if C4 is 0/empty, otherwise calculates hours)
        cell_hours = ws_sum.cell(row=current_row, column=5)
        cell_hours.value = f'=IF(C{current_row}=0, 0, (D{current_row}-C{current_row})*24)'
        cell_hours.number_format = '0.00'
        cell_hours.alignment = right_align
        cell_hours.font = body_font
        
        # Col F: Is Workday
        cell_workday = ws_sum.cell(row=current_row, column=6)
        cell_workday.value = f'=IF(E{current_row}>0, 1, 0)'
        cell_workday.number_format = '0'
        cell_workday.alignment = center_align
        cell_workday.font = body_font
        
        # Col G: Week #
        cell_week = ws_sum.cell(row=current_row, column=7)
        cell_week.value = f'=WEEKNUM(A{current_row}, 2)' # Monday start
        cell_week.number_format = '0'
        cell_week.alignment = center_align
        cell_week.font = body_font
        
        # Borders and alternating fill
        for col in range(1, 8):
            c = ws_sum.cell(row=current_row, column=col)
            c.border = thin_border
            if current_row % 2 == 1:
                c.fill = zebra_fill
                
        current_row += 1
        
    # ----------------------------------------------------
    # Sheet 1: Weekly Summary Side Table
    # ----------------------------------------------------
    weekly_headers = [
        ("Week", "I3", center_align),
        ("Total Hours", "J3", right_align),
        ("Work Days", "K3", center_align)
    ]
    
    for label, cell_ref, align in weekly_headers:
        cell = ws_sum[cell_ref]
        cell.value = label
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align
        cell.border = thin_border
        
    # Weeks 23, 24, 25
    week_rows = [
        (23, 4),
        (24, 5),
        (25, 6)
    ]
    
    for wk, r in week_rows:
        cell_wk = ws_sum.cell(row=r, column=9)
        cell_wk.value = wk
        cell_wk.number_format = '"Week "0'
        cell_wk.alignment = center_align
        cell_wk.font = bold_body_font
        cell_wk.border = thin_border
        
        cell_tot_hrs = ws_sum.cell(row=r, column=10)
        cell_tot_hrs.value = f'=SUMIFS($E$4:$E$100, $G$4:$G$100, I{r})'
        cell_tot_hrs.number_format = '0.00'
        cell_tot_hrs.alignment = right_align
        cell_tot_hrs.font = body_font
        cell_tot_hrs.border = thin_border
        
        cell_tot_days = ws_sum.cell(row=r, column=11)
        cell_tot_days.value = f'=SUMIFS($F$4:$F$100, $G$4:$G$100, I{r})'
        cell_tot_days.number_format = '0'
        cell_tot_days.alignment = center_align
        cell_tot_days.font = body_font
        cell_tot_days.border = thin_border
        
    # Grand Total (Row 7)
    ws_sum.cell(row=7, column=9, value="Grand Total").font = total_label_font
    ws_sum.cell(row=7, column=9).alignment = center_align
    ws_sum.cell(row=7, column=9).border = total_border
    ws_sum.cell(row=7, column=9).fill = total_highlight_fill
    
    cell_grand_hrs = ws_sum.cell(row=7, column=10)
    cell_grand_hrs.value = "=SUM(J4:J6)"
    cell_grand_hrs.number_format = '0.00'
    cell_grand_hrs.font = bold_body_font
    cell_grand_hrs.alignment = right_align
    cell_grand_hrs.border = total_border
    cell_grand_hrs.fill = total_highlight_fill
    
    cell_grand_days = ws_sum.cell(row=7, column=11)
    cell_grand_days.value = "=SUM(K4:K6)"
    cell_grand_days.number_format = '0'
    cell_grand_days.font = bold_body_font
    cell_grand_days.alignment = center_align
    cell_grand_days.border = total_border
    cell_grand_days.fill = total_highlight_fill
    ws_sum.row_dimensions[7].height = 22
    
    ws_sum.merge_cells("I9:K11")
    note_cell = ws_sum["I9"]
    note_cell.value = "Note:\nTotal work hours per week represent elapsed span from first file activity to last file activity. Week numbers start on Mondays (standard ISO)."
    note_cell.font = gray_italics_font
    note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # ----------------------------------------------------
    # Sheet 2: Activity
    # ----------------------------------------------------
    ws_act = wb.create_sheet(title="Activity")
    ws_act.views.sheetView[0].showGridLines = True
    
    act_headers = [
        ("Work Date", 1, center_align),
        ("Folder Name", 2, left_align),
        ("Folder Path", 3, left_align),
        ("Earliest Action", 4, center_align),
        ("Latest Action", 5, center_align),
        ("File Modifications", 6, center_align)
    ]
    
    ws_act.row_dimensions[1].height = 24
    for label, col, align in act_headers:
        cell = ws_act.cell(row=1, column=col)
        cell.value = label
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align
        cell.border = thin_border
        
    act_row = 2
    for r in aggregated_data:
        ws_act.row_dimensions[act_row].height = 18
        
        # Col A: Work Date
        c_date = ws_act.cell(row=act_row, column=1)
        c_date.value = datetime.datetime.combine(r['work_date'], datetime.time.min)
        c_date.number_format = 'yyyy-mm-dd'
        c_date.alignment = center_align
        c_date.font = body_font
        
        # Col B: Folder Name
        c_name = ws_act.cell(row=act_row, column=2)
        c_name.value = r['folder_name']
        c_name.alignment = left_align
        c_name.font = body_font
        
        # Col C: Folder Path
        c_path = ws_act.cell(row=act_row, column=3)
        c_path.value = r['folder_path']
        c_path.alignment = left_align
        c_path.font = body_font
        
        # Col D: Earliest Action
        c_earliest = ws_act.cell(row=act_row, column=4)
        c_earliest.value = r['earliest_time']
        c_earliest.number_format = 'yyyy-mm-dd hh:mm AM/PM'
        c_earliest.alignment = center_align
        c_earliest.font = body_font
        
        # Col E: Latest Action
        c_latest = ws_act.cell(row=act_row, column=5)
        c_latest.value = r['latest_time']
        c_latest.number_format = 'yyyy-mm-dd hh:mm AM/PM'
        c_latest.alignment = center_align
        c_latest.font = body_font
        
        # Col F: File Modifications
        c_edits = ws_act.cell(row=act_row, column=6)
        c_edits.value = r['edits_count']
        c_edits.number_format = '0'
        c_edits.alignment = center_align
        c_edits.font = body_font
        
        # Borders and alternating fill
        for col in range(1, 7):
            c = ws_act.cell(row=act_row, column=col)
            c.border = thin_border
            if act_row % 2 == 1:
                c.fill = zebra_fill
                
        act_row += 1
        
    # ----------------------------------------------------
    # Sheet 2: Folder Summary Side Table (Columns H:I)
    # ----------------------------------------------------
    unique_folders = sorted(list(set(r['folder_name'] for r in aggregated_data)))
    
    # Headers
    fold_headers = [
        ("Folder Name Summary", "H1", left_align),
        ("Total Edits", "I1", center_align)
    ]
    
    for label, cell_ref, align in fold_headers:
        cell = ws_act[cell_ref]
        cell.value = label
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align
        cell.border = thin_border
        
    # Populate Folder Summary table
    fold_row = 2
    for folder in unique_folders:
        ws_act.row_dimensions[fold_row].height = 18
        
        # Col H: Folder Name
        c_fold = ws_act.cell(row=fold_row, column=8)
        c_fold.value = folder
        c_fold.alignment = left_align
        c_fold.font = bold_body_font
        c_fold.border = thin_border
        
        # Col I: Total Edits in Folder (Formula: sum of Column F matching folder)
        c_f_tot_edits = ws_act.cell(row=fold_row, column=9)
        c_f_tot_edits.value = f'=SUMIF($B$2:$B$5000, H{fold_row}, $F$2:$F$5000)'
        c_f_tot_edits.number_format = '0'
        c_f_tot_edits.alignment = center_align
        c_f_tot_edits.font = body_font
        c_f_tot_edits.border = thin_border
        
        if fold_row % 2 == 1:
            c_fold.fill = zebra_fill
            c_f_tot_edits.fill = zebra_fill
            
        fold_row += 1
        
    # Folder Summary Grand Total Row
    ws_act.cell(row=fold_row, column=8, value="Grand Total").font = total_label_font
    ws_act.cell(row=fold_row, column=8).alignment = left_align
    ws_act.cell(row=fold_row, column=8).border = total_border
    ws_act.cell(row=fold_row, column=8).fill = total_highlight_fill
    
    c_f_tot_grand_edits = ws_act.cell(row=fold_row, column=9)
    c_f_tot_grand_edits.value = f'=SUM(I2:I{fold_row-1})'
    c_f_tot_grand_edits.number_format = '0'
    c_f_tot_grand_edits.font = bold_body_font
    c_f_tot_grand_edits.alignment = center_align
    c_f_tot_grand_edits.border = total_border
    c_f_tot_grand_edits.fill = total_highlight_fill
    ws_act.row_dimensions[fold_row].height = 22
    
    # Auto-fit columns
    for ws in [ws_sum, ws_act]:
        for col in ws.columns:
            if ws.title == "Summary" and col[0].column in [1, 2, 3, 4, 5, 6, 7] and (col[0].row == 1 or col[0].row == 2):
                max_len = 0
                for cell in col[2:]:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            else:
                max_len = 0
                for cell in col:
                    if cell.value:
                        val_str = str(cell.value)
                        if not val_str.startswith('='):
                            max_len = max(max_len, len(val_str))
                col_letter = get_column_letter(col[0].column)
                if ws.title == "Summary" and col_letter in ['C', 'D']:
                    ws.column_dimensions[col_letter].width = 24
                elif ws.title == "Activity" and col_letter == 'C':
                    ws.column_dimensions[col_letter].width = 45
                elif ws.title == "Activity" and col_letter == 'H':
                    ws.column_dimensions[col_letter].width = 30
                else:
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
                    
    wb.save(output_filename)
    print(f"Spreadsheet generated successfully. Saved to: {output_filename}")

def main():
    parser = argparse.ArgumentParser(description="Generate Texas 2036 Timesheet.")
    parser.add_argument("--start", default=DEFAULT_START_DATE, help="Start Date (YYYY-MM-DD)")
    parser.add_argument("--end", default=DEFAULT_END_DATE, help="End Date (YYYY-MM-DD)")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_FILE, help="Output Excel filename")
    args = parser.parse_args()
    
    try:
        start_date = datetime.datetime.strptime(args.start, "%Y-%m-%d").date()
        end_date = datetime.datetime.strptime(args.end, "%Y-%m-%d").date()
    except ValueError:
        print("Error: Dates must be in YYYY-MM-DD format.")
        sys.exit(1)
        
    if start_date > end_date:
        print("Error: Start date cannot be after end date.")
        sys.exit(1)
        
    # Scan windows
    start_dt = datetime.datetime.combine(start_date, datetime.time(2, 0, 0))
    end_dt = datetime.datetime.combine(end_date + datetime.timedelta(days=1), datetime.time(2, 0, 0))
    
    print(f"Time Sheet Period: {start_date} to {end_date}")
    print(f"Scan Window: {start_dt} to {end_dt}")
    print("----------------------------------------")
    
    # 1. Gather file modifications
    file_records = scan_files(start_dt, end_dt)
    
    if not file_records:
        print("No activity records found in the specified timeframe.")
        sys.exit(0)
        
    # 2. Group and aggregate
    aggregated_data = aggregate_activity(file_records)
    
    # 3. Generate Excel
    generate_excel(aggregated_data, start_date, end_date, args.output)
    
if __name__ == "__main__":
    main()
